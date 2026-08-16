#!/usr/bin/env python3
"""Gera uma planilha de plano e diário a partir de uma tabela Markdown.

A tabela deve ter as colunas: Encontro, Cap., Data, Modalidade, Unidade,
Tipo, Tema, Resultados, Conteúdos, Estratégia, Marco, Bloco 1 e Bloco 2.
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


EXPECTED = [
    "Encontro", "Cap.", "Data", "Modalidade", "Unidade", "Tipo", "Tema",
    "Resultados", "Conteúdos", "Estratégia", "Marco", "Bloco 1", "Bloco 2",
]

DIARY = [
    "Data realizada", "Conteúdo ministrado", "Metodologia/recursos usados",
    "Evidências/links do Classroom", "Presença e observações",
    "Pendências/próxima aula", "Status no Classroom", "Atualizado em",
]


def clean(cell: str) -> str:
    return re.sub(r"\s+", " ", cell.strip().replace("<br>", "\n"))


def parse_table(path: Path) -> list[list[str]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    for index, line in enumerate(lines):
        if line.strip().startswith("| Encontro |"):
            header = [clean(x) for x in line.strip().strip("|").split("|")]
            if header != EXPECTED:
                raise ValueError(f"Cabeçalho inesperado: {header!r}")
            rows: list[list[str]] = []
            for item in lines[index + 2 :]:
                if not item.strip().startswith("|"):
                    break
                row = [clean(x) for x in item.strip().strip("|").split("|")]
                if len(row) != len(EXPECTED):
                    raise ValueError(f"Linha com {len(row)} colunas: {item}")
                rows.append(row)
            if not rows:
                raise ValueError("Tabela encontrada sem encontros")
            return rows
    raise ValueError("Tabela canônica não encontrada")


def build(plan: Path, output: Path, discipline: str) -> None:
    rows = parse_table(plan)
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano e diário"
    headers = EXPECTED + DIARY
    ws.append(headers)

    for row in rows:
        ws.append(row + [""] * len(DIARY))

    navy = "17365D"
    blue = "D9EAF7"
    green = "E2F0D9"
    orange = "FCE4D6"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7C9D6")

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for idx, cell in enumerate(row, start=1):
            cell.fill = PatternFill("solid", fgColor=blue if idx <= len(EXPECTED) else green)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
        if row[5].value and row[5].value.lower() != "aula":
            for cell in row:
                cell.fill = PatternFill("solid", fgColor=orange)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 42
    for idx in range(2, ws.max_row + 1):
        ws.row_dimensions[idx].height = 78

    widths = [10, 8, 12, 15, 10, 16, 28, 36, 42, 34, 24, 36, 36,
              15, 42, 36, 34, 30, 34, 20, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = width

    status_col = len(EXPECTED) + 7
    dv = DataValidation(type="list", formula1='"Não publicado,Rascunho,Publicado,Atualizado"')
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(2, status_col).coordinate}:{ws.cell(ws.max_row, status_col).coordinate}")
    ws.conditional_formatting.add(
        f"A2:{ws.cell(ws.max_row, len(headers)).coordinate}",
        FormulaRule(formula=[f'$F2="Avaliação"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    info = wb.create_sheet("Como usar")
    info.append([f"{discipline} — plano e diário 2026.2"])
    info.append(["As colunas azuis preservam o planejado; as verdes registram o realizado."])
    info.append(["Não apague o planejado quando houver mudança. Registre a diferença e a reposição."])
    info.append(["Atualize evidências/links somente após publicar o material no Classroom."])
    info.column_dimensions["A"].width = 110
    for cell in info[1]:
        cell.font = Font(bold=True, color=white)
        cell.fill = PatternFill("solid", fgColor=navy)

    wb.properties.title = f"{discipline} — Plano e diário 2026.2"
    wb.properties.creator = "Cayo Oliveira"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("plan", type=Path)
    parser.add_argument("output", type=Path)
    parser.add_argument("--discipline", required=True)
    args = parser.parse_args()
    build(args.plan, args.output, args.discipline)


if __name__ == "__main__":
    main()
